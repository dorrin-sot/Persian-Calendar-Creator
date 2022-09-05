import pathlib
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from calendar_model import generate_days, PERSIAN_MONTHS, GREGORIAN_MONTHS, ISLAMIC_MONTHS, WEEKDAYS


def get_column_name_from_num(weekday: int) -> str:
    return "abcdefghijklmn".upper()[weekday]


def persian(num: int) -> str:
    return "".join(["۰۱۲۳۴۵۶۷۸۹"[int(c)] for c in str(num)])


def arabic(num: int) -> str:
    return "".join(["٠١٢٣٤٥٦٧٨٩"[int(c)] for c in str(num)])


days = generate_days(int(input("Jalali year: ")))
months = []
for day in days:
    if day.persian[2] == 1:
        months.append([])
    months[day.persian[1] - 1].append(day)

wb = Workbook()
wb.remove(wb.active)  # delete initial sheet

# colors
gunmetal = "2C363F"
blush = "E53245"
ivory = "F2F5EA"
timberwolf = "D6DBD2"
laurel_green = "BBC7A4"


def style_persian_month_title(ws: Worksheet):
    ws.row_dimensions[1].height = 50
    for cols in ws.iter_cols(min_col=1, max_col=14, min_row=1, max_row=1):
        cell = cols[0]
        cell.fill = PatternFill(fill_type="solid", start_color=laurel_green, end_color=laurel_green)
        cell.font = Font(name="Cinema", size=40, bold=True, color=gunmetal)


def style_other_month_titles(ws: Worksheet):
    ws.row_dimensions[2].height = 31
    ws.merge_cells("A2:G2")
    ws.merge_cells("H2:N2")
    for cols in ws.iter_cols(min_col=1, max_col=14, min_row=2, max_row=2):
        cell = cols[0]
        cell.fill = PatternFill(fill_type="solid", start_color=laurel_green, end_color=laurel_green)
        cell.font = Font(name="Cinema", size=25, bold=True, color=gunmetal)

    islamic_cell = ws["A2"]
    islamic_cell.font = Font(name="Calibri", size=18)
    islamic_cell.alignment = Alignment(horizontal="right", vertical="center")

    gregorian_cell = ws["H2"]
    gregorian_cell.font = Font(name="Calibri", size=18)
    gregorian_cell.alignment = Alignment(horizontal="left", vertical="center")


def style_weekdays(ws: Worksheet):
    ws.row_dimensions[2].height = 35
    for cols in ws.iter_cols(min_col=1, max_col=14, min_row=3, max_row=3):
        cell = cols[0]
        cell.fill = PatternFill(fill_type="solid", start_color=timberwolf, end_color=timberwolf)
        cell.font = Font(name="Far.Domrol", size=18, bold=True, color=gunmetal)


def style_day_big(cell):
    cell.font = Font(name="Calibri", size=35, bold=True, color=cell.font.color)


def style_day_normal(cell):
    cell_str = str(cell.value)
    cell.font = Font(
        name="Calibri" if re.search("[0-9]+( \D+)?", cell_str) else "Calibri",
        size=9 if " " in cell_str else 17,
        color=cell.font.color
    )


# noinspection PyShadowingNames
def style_days(ws: Worksheet):
    for col in "abcdefghijklmn".upper():
        ws.column_dimensions[col].width = 8.5
    for i in range(4, 16):
        ws.row_dimensions[i].height = 30
    for rows in ws.iter_rows(min_row=4, max_row=15, min_col=1, max_col=14):
        for cell in rows:
            cell.fill = PatternFill(patternType="solid", start_color=ivory, end_color=ivory)
            if cell.column % 2 == 1:
                style_day_big(cell)
            else:
                style_day_normal(cell)


def set_month_names(ws: Worksheet, persian_month):
    start, end, prev = None, None, None
    for day in days:
        if start is None and day.persian[1] == persian_month:
            start = day
        if start is not None:
            if persian_month == 12:
                end = days[-1]
            if day.persian[1] == persian_month + 1:
                end = prev
            if end is not None:
                break
        prev = day
    ws["A1"] = PERSIAN_MONTHS[monthNum]
    ws["A2"] = str.format(
        "  {}{} - {}{}",
        ISLAMIC_MONTHS[start.islamic[1] - 1],
        f" {arabic(start.islamic[0])}" if start.islamic[0] != end.islamic[0] else "",
        ISLAMIC_MONTHS[end.islamic[1] - 1],
        f" {arabic(end.islamic[0])}" if start.islamic[0] != end.islamic[0] else "",
    )
    ws["H2"] = str.format(
        "  {}{} - {}{}",
        GREGORIAN_MONTHS[start.gregorian[1] - 1],
        f" {start.gregorian[0]}" if start.gregorian[0] != end.gregorian[0] else "",
        GREGORIAN_MONTHS[end.gregorian[1] - 1],
        f" {end.gregorian[0]}" if start.gregorian[0] != end.gregorian[0] else "",
    )


for monthNum, month in enumerate(months):
    monthName = PERSIAN_MONTHS[monthNum]
    wb.create_sheet(monthName)
    wb.active = wb[monthName]
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    Worksheet.set_printer_settings(ws, paper_size=9, orientation='landscape')

    ws.row_dimensions[1].height = 50

    # add days
    prev = None
    for day in month:
        # print(day)
        # persian day
        start = f"{get_column_name_from_num(2 * day.week_day)}{2 * day.week_num + 2}"
        end = f"{get_column_name_from_num(2 * day.week_day)}{2 * day.week_num + 3}"
        ws.merge_cells(f"{start}:{end}")
        ws[start] = persian(day.persian[2])
        if day.is_holiday:
            ws[start].font = Font(color=blush)

        # gregorian day
        start = f"{get_column_name_from_num(2 * day.week_day + 1)}{2 * day.week_num + 2}"
        ws[start] = str.format("{}{}",
                               int(day.gregorian[2]),
                               f" {GREGORIAN_MONTHS[day.gregorian[1] - 1]}" if prev is not None and day.gregorian[1] != prev.gregorian[1] else "")
        if day.is_holiday:
            ws[start].font = Font(color=blush)

        # islamic day
        start = f"{get_column_name_from_num(2 * day.week_day + 1)}{2 * day.week_num + 3}"
        ws[start] = day.islamic[2]
        ws[start] = str.format("{}{}",
                               arabic(int(day.islamic[2])),
                               f" {ISLAMIC_MONTHS[day.islamic[1] - 1]}" if prev is not None and day.islamic[1] != prev.islamic[1] else "")
        if day.is_holiday:
            ws[start].font = Font(color=blush)

        prev = day

    for i in range(21)[1:]:
        for j in "abcdefghijklmno".upper():
            ws[f"{j}{i}"].alignment = Alignment(horizontal="center", vertical="center", readingOrder=2)

    # set month name titles
    ws.merge_cells(f"{get_column_name_from_num(0)}1:{get_column_name_from_num(13)}1")
    set_month_names(ws, monthNum + 1)
    style_persian_month_title(ws)
    style_other_month_titles(ws)
    style_weekdays(ws)
    style_days(ws)

    # set weekdays
    for i in range(7):
        start = f"{get_column_name_from_num(2 * i)}3"
        end = f"{get_column_name_from_num(2 * i + 1)}3"
        ws.merge_cells(f"{start}:{end}")
        ws[start] = WEEKDAYS[i]

# Save the file
wb.create_sheet("New Sheet")
wb.active = wb["New Sheet"]
wb.save("calendar.xlsx")

print(f"Done! Your Excel file is saved in {pathlib.Path(__file__).parent.resolve()}/calendar.xslx")
